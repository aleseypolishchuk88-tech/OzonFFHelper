"""Генератор листа «Оприходование» (XLSX) для ФФ из состава поставки Ozon.

Формат повторяет шаблон заказчика («Оприходования Иванова»):
  строки 1–5  — шапка-метаданные (тип документа, № и дата, поклажедатель, ИНН);
  строка 6    — пустая;
  строка 7    — заголовки таблицы;
  строки 8+   — по одной на товар: ШК(OZN) | Наименование | Размер | Цвет | Кол-во | Цена | Примечание.
"""
from __future__ import annotations

import io
import re
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from core.models import SupplyContentItem

# Значения по умолчанию из шаблона заказчика.
DEFAULT_COUNTERPARTY = "Полищук Алексей Анатольевич"
DEFAULT_INN = "111701352298"
DEFAULT_PRICE = 3000
DEFAULT_NOTE = ""  # примечание по умолчанию пустое (заполняется вручную при нужде)

# Оформление — точно по эталонному шаблону заказчика.
_FONT = Font(name="Calibri", size=11)
_TITLE_FONT = Font(name="Calibri", size=21, bold=True)
_THIN = Side(style="thin", color="FF000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="FFF3F0E5")  # кремовая шапка как в оригинале
_AL_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
_AL_CENTER = Alignment(horizontal="center", vertical="top")
# Колонки таблицы и их ширины (D — по умолчанию, как в оригинале).
_COL_WIDTHS = {"A": 26.4, "B": 77.4, "C": 15.3, "E": 11.1, "F": 12.6, "G": 42.0}

TABLE_HEADERS = [
    "ШК",
    "Наименование товара",
    "Размер",
    "Цвет",
    "Кол-во ",  # пробел в конце — как в шаблоне ФФ
    "Цена, руб.",
    "Примечание (не обяз.)",
]

# Цвета для извлечения из названия (нижний регистр -> отображаемое).
_COLORS = {
    "белый": "Белый",
    "белая": "Белый",
    "белые": "Белый",
    "черный": "Черный",
    "чёрный": "Черный",
    "черная": "Черный",
    "чёрная": "Черный",
    "серый": "Серый",
    "серебристый": "Серебристый",
    "золотой": "Золотой",
    "золотистый": "Золотистый",
    "бронзовый": "Бронзовый",
    "бежевый": "Бежевый",
    "коричневый": "Коричневый",
    "синий": "Синий",
    "зеленый": "Зеленый",
    "зелёный": "Зеленый",
    "красный": "Красный",
}


def extract_color(name: str | None) -> str:
    """Попытаться вытащить цвет из названия товара. Пусто, если не нашли."""
    if not name:
        return ""
    words = re.findall(r"[А-Яа-яЁё]+", name)
    for w in reversed(words):  # цвет обычно в конце названия
        c = _COLORS.get(w.lower())
        if c:
            return c
    return ""


def build_workbook(
    items: list[SupplyContentItem],
    *,
    doc_number: str = "1",
    doc_date: date | None = None,
    counterparty: str = DEFAULT_COUNTERPARTY,
    inn: str = DEFAULT_INN,
    price: int = DEFAULT_PRICE,
    note: str = DEFAULT_NOTE,
) -> Workbook:
    """Собрать XLSX-книгу «Оприходование»."""
    doc_date = doc_date or date.today()
    wb = Workbook()
    ws = wb.active
    ws.title = "Основной"  # как в эталонном шаблоне

    ws["A1"] = "Перемещение на ответственное хранение"
    ws["A1"].font = _TITLE_FONT

    meta = [
        ("Номер входящего документа", doc_number),
        ("Дата входящего документа", doc_date),
        ("Наименование поклажедатель", counterparty),
        ("ИНН поклажедатель", inn),
    ]
    for i, (key, val) in enumerate(meta, start=2):
        kc = ws.cell(row=i, column=1, value=key)
        vc = ws.cell(row=i, column=2, value=val)
        for c in (kc, vc):
            c.font = _FONT
            c.border = _BORDER
            c.alignment = _AL_LEFT
    ws["B3"].number_format = "mm-dd-yy"  # дата как в оригинале
    # строка 6 — пустая

    # Шапка таблицы (строка 7): кремовая заливка, не жирная, границы.
    for col, title in enumerate(TABLE_HEADERS, start=1):
        cell = ws.cell(row=7, column=col, value=title)
        cell.font = _FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _AL_CENTER if col == 5 else _AL_LEFT
        cell.border = _BORDER

    # Данные: по центру только «Кол-во» (колонка 5), остальное — слева/сверху.
    row = 8
    for it in items:
        values = [
            it.ozn_code,
            it.name or "",
            "",  # Размер
            extract_color(it.name),
            it.quantity,
            price,
            note,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = _FONT
            cell.border = _BORDER
            cell.alignment = _AL_CENTER if col == 5 else _AL_LEFT
        row += 1

    for col, w in _COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    return wb


def build_bytes(items: list[SupplyContentItem], **kwargs) -> bytes:
    """XLSX в виде bytes (для кнопки скачивания в Streamlit)."""
    wb = build_workbook(items, **kwargs)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_to_file(items: list[SupplyContentItem], path: str | Path, **kwargs) -> Path:
    """Сохранить XLSX на диск, вернуть путь."""
    path = Path(path)
    build_workbook(items, **kwargs).save(path)
    return path


def suggest_filename(order_number: str, warehouse: str | None = None) -> str:
    """Имя файла: Оприходование_<склад>_<номер>_<дата>.xlsx."""
    wh = re.sub(r"[^A-Za-zА-Яа-я0-9]+", "_", (warehouse or "").strip()).strip("_")
    wh = f"_{wh}" if wh else ""
    return f"Оприходование{wh}_{order_number}_{date.today():%Y-%m-%d}.xlsx"
